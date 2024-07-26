import openpyxl as opxl
import dataclasses as dtcls

@dtcls.dataclass
class Report:
    model_name: str = None
    loss_name: str = None
    train_count: int = None
    valid_count: int = None
    batch_size: int = None
    epochs: int = None
    loss_value: float = None
    accuracy_value: float = None
    folder: str = None
    comment: str = None

class ExcelParser:
    FILE_NAME = 'report.xlsx'
    def _init_table(self) -> None:
        ws = self.wb.active
        ws.append([None] + list(Report.__match_args__))
        self.wb.save(self.path)

    def __init__(self, path='') -> None:
        self.wb = opxl.Workbook()
        import os
        self.path = os.path.join(path, self.FILE_NAME)

        if not os.path.exists(self.path): self._init_table()

        self.wb = opxl.load_workbook(self.path)

    def __exit__(self) -> None:
        self.wb.close()

    def save(self, data: Report) -> None:
        ws = self.wb.active
        ws.append([ws.max_row] + list(dtcls.asdict(data).values()))
        self.wb.save(self.path)

class Logger:
    def __init__(self, path: str) -> None:
        import os
        if not os.path.exists(path): os.makedirs(path)

        self.parser = ExcelParser(path)
        self.path = path
    
    def __call__(self, name: str) -> str:
        import time, os
        self.folder_name = f'{time.strftime("%Y-%m-%d_%H-%M-%S", time.localtime())}_{name}'
        self.path = os.path.join(self.path, self.folder_name)
        if not os.path.exists(self.path): os.makedirs(self.path)
        return self.path

    def save_results(self, report: Report) -> None:
        report.folder = self.folder_name
        self.parser.save(report)
    
    def save_plot(self, data: any, name: str) -> None:
        import matplotlib.pyplot as plt
        plt.plot(data[0])
        plt.plot(data[1])
        import os
        graphic_path = os.path.join(self.path, name)
        plt.savefig(graphic_path)
        plt.close()

    def save_predict(self, data: any, size=4) -> None:
        import matplotlib.pyplot as plt
        plt.figure(figsize=(10, 4))
        
        for x, images in enumerate(data):
            for i, image in enumerate(images[:size][:8]):
                plt.subplot(len(data), size, i + x * size + 1)
                plt.imshow(image, cmap='gray')
                plt.axis("off")
        
        import os
        image_path = os.path.join(self.path, 'image')
        plt.savefig(image_path)
        plt.close()